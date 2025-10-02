import pandas as pd
import numpy as np
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
import logging

class ReportGenerator:
    """Генератор расширенных отчетов в Excel формате"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def generate_report(self, results: List[Dict[str, Any]], 
                       output_path: str, 
                       include_charts: bool = True) -> pd.DataFrame:
        """Генерация полного отчета"""
        
        # Создание основного DataFrame
        df = self._create_dataframe(results)
        
        # Сохранение в Excel с форматированием
        self._save_excel_report(df, output_path, include_charts)
        
        # Генерация сводной статистики
        self._generate_summary_statistics(df, output_path)
        
        return df
    
    def _create_dataframe(self, results: List[Dict[str, Any]]) -> pd.DataFrame:
        """Создание DataFrame с результатами"""
        
        # Базовые колонки
        base_columns = [
            'path_to_study', 'study_uid', 'series_uid', 
            'probability_of_pathology', 'pathology', 
            'processing_status', 'time_of_processing'
        ]
        
        
        # Создание DataFrame
        df_data = []
        for result in results:
            row = {}
            
            # Базовые поля
            for col in base_columns:
                row[col] = result.get(col, '')
            
            
            df_data.append(row)
        
        return pd.DataFrame(df_data)
    
    def _save_excel_report(self, df: pd.DataFrame, output_path: str, include_charts: bool):
        """Сохранение отчета в Excel с форматированием"""
        
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Основной лист с результатами
            df.to_excel(writer, sheet_name='Results', index=False)
            
            # Лист со статистикой
            self._add_statistics_sheet(writer, df)
            
            # Лист с ошибками если есть
            if not df[df['processing_status'] == 'Success'].empty:
                self._add_errors_sheet(writer, df)
            
            # Форматирование
            workbook = writer.book
            self._format_workbook(workbook, df)
            
            # Графики если требуется
            if include_charts:
                self._add_charts(workbook, df)
    
    def _add_statistics_sheet(self, writer, df: pd.DataFrame):
        """Добавление листа со статистикой"""
        
        statistics_data = []
        
        # Основная статистика
        total_studies = len(df)
        successful = len(df[df['processing_status'] == 'Success'])
        failed = len(df[df['processing_status'] != 'Success'])
        
        statistics_data.extend([
            ['Metric', 'Value'],
            ['Total Studies', total_studies],
            ['Successful Processing', successful],
            ['Failed Processing', failed],
            ['Success Rate', f'{(successful/total_studies)*100:.1f}%' if total_studies > 0 else '0%'],
            ['', ''],
        ])
        
        # Статистика по патологиям если есть успешные обработки
        if successful > 0:
            pathology_stats = df[df['processing_status'] == 'Success']['pathology'].value_counts()
            normal_count = pathology_stats.get(0, 0)
            pathology_count = pathology_stats.get(1, 0)
            
            statistics_data.extend([
                ['Pathology Statistics', ''],
                ['Normal Studies', normal_count],
                ['Studies with Pathology', pathology_count],
                ['Pathology Rate', f'{(pathology_count/successful)*100:.1f}%' if successful > 0 else '0%'],
                ['', ''],
            ])
        
        # Временная статистика
        if 'time_of_processing' in df.columns:
            time_stats = df['time_of_processing'].describe()
            statistics_data.extend([
                ['Processing Time Statistics (seconds)', ''],
                ['Mean', f"{time_stats.get('mean', 0):.2f}"],
                ['Std', f"{time_stats.get('std', 0):.2f}"],
                ['Min', f"{time_stats.get('min', 0):.2f}"],
                ['Max', f"{time_stats.get('max', 0):.2f}"],
            ])
        
        # Создание DataFrame и сохранение
        stats_df = pd.DataFrame(statistics_data[1:], columns=statistics_data[0])
        stats_df.to_excel(writer, sheet_name='Statistics', index=False)
    
    def _add_errors_sheet(self, writer, df: pd.DataFrame):
        """Добавление листа с ошибками"""
        
        error_data = []
        failed_studies = df[df['processing_status'] != 'Success']
        
        for _, row in failed_studies.iterrows():
            error_data.append({
                'Study Path': row['path_to_study'],
                'Error Message': row.get('error_message', 'Unknown error'),
                'Study UID': row.get('study_uid', ''),
                'Series UID': row.get('series_uid', '')
            })
        
        if error_data:
            error_df = pd.DataFrame(error_data)
            error_df.to_excel(writer, sheet_name='Errors', index=False)
    
    def _format_workbook(self, workbook: Workbook, df: pd.DataFrame):
        """Форматирование рабочей книги"""
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Форматирование листа Results
        if 'Results' in workbook.sheetnames:
            ws = workbook['Results']
            
            # Заголовки
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Форматирование данных
            for row in range(2, len(df) + 2):
                # Статус обработки
                status_cell = ws[f'F{row}']  # processing_status
                if status_cell.value == 'Success':
                    status_cell.fill = success_fill
                else:
                    status_cell.fill = error_fill
                
                # Вероятность патологии
                prob_cell = ws[f'D{row}']  # probability_of_pathology
                if isinstance(prob_cell.value, (int, float)):
                    prob_cell.number_format = '0.000'
                    
                    # Цветовое кодирование вероятности
                    if prob_cell.value > 0.7:
                        prob_cell.fill = PatternFill(start_color="FF9999", fill_type="solid")
                    elif prob_cell.value > 0.3:
                        prob_cell.fill = PatternFill(start_color="FFFF99", fill_type="solid")
                    else:
                        prob_cell.fill = PatternFill(start_color="99FF99", fill_type="solid")
                
                # Время обработки
                time_cell = ws[f'G{row}']  # time_of_processing
                if isinstance(time_cell.value, (int, float)):
                    time_cell.number_format = '0.000'
            
            # Автоматическая ширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_charts(self, workbook: Workbook, df: pd.DataFrame):
        """Добавление графиков в отчет"""
        
        if 'Results' not in workbook.sheetnames:
            return
        
        ws = workbook['Results']
        chart_ws = workbook.create_sheet(title="Charts")
        
        # График распределения вероятностей
        if 'probability_of_pathology' in df.columns:
            success_df = df[df['processing_status'] == 'Success']
            if not success_df.empty:
                chart1 = LineChart()
                chart1.title = "Distribution of Pathology Probabilities"
                chart1.style = 13
                
                # Подготовка данных для графика
                probabilities = success_df['probability_of_pathology'].values
                hist, bins = np.histogram(probabilities, bins=20, range=(0, 1))
                
                # Добавление данных
                chart_data = Reference(ws, min_col=4, min_row=2, 
                                     max_row=len(hist)+1, max_col=4)
                chart1.add_data(chart_data, titles_from_data=True)
                
                chart_ws.add_chart(chart1, "A1")
        
        # График времени обработки
        if 'time_of_processing' in df.columns:
            chart2 = LineChart()
            chart2.title = "Processing Time Distribution"
            chart2.style = 13
            
            times = df['time_of_processing'].values
            hist_times, bins_times = np.histogram(times, bins=15)
            
            # Добавление данных времени обработки
            chart_data_time = Reference(ws, min_col=7, min_row=2, 
                                      max_row=len(hist_times)+1, max_col=7)
            chart2.add_data(chart_data_time, titles_from_data=True)
            
            chart_ws.add_chart(chart2, "A20")
    
    def _generate_summary_statistics(self, df: pd.DataFrame, output_path: str):
        """Генерация текстовой сводки"""
        
        summary_path = output_path.replace('.xlsx', '_summary.txt')
        
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write("CT Classification Report Summary\n")
            f.write("=" * 50 + "\n\n")
            
            f.write(f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total studies processed: {len(df)}\n")
            
            successful = len(df[df['processing_status'] == 'Success'])
            f.write(f"Successfully processed: {successful}\n")
            f.write(f"Success rate: {(successful/len(df))*100:.1f}%\n\n")
            
            if successful > 0:
                pathology_count = len(df[(df['processing_status'] == 'Success') & 
                                       (df['pathology'] == 1)])
                f.write(f"Studies with pathology: {pathology_count}\n")
                f.write(f"Pathology detection rate: {(pathology_count/successful)*100:.1f}%\n\n")
            
            # Среднее время обработки
            if 'time_of_processing' in df.columns:
                avg_time = df['time_of_processing'].mean()
                f.write(f"Average processing time: {avg_time:.2f} seconds\n")
            
            f.write("\nDetailed report available in: " + os.path.basename(output_path))